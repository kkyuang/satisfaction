import pandas as pd
import urllib.request
import json
import ssl
from openpyxl import load_workbook

#네이버 API
client_id = "7yE3WNUE_IdJTOuoOUjX"
client_secret = "M189DprmrK"


#엑셀 읽기
workbook = load_workbook('memelist.xlsx')

# 시트 선택
sheet = workbook['basicList']

# 데이터 입력 받기
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행은 헤더이므로 건너뜀
    data.append(row)

# 데이터를 이용하여 body 생성
body_data = []
for row in data:
    group = {
        "groupName": row[0],
        "keywords": [kw.strip() for kw in row[1].split(',')]
    }
    body_data.append(group)

def getData(n, m):

    # body JSON 생성
    body = {
        "startDate": "2022-01-01",
        "endDate": "2023-07-07",
        "timeUnit": "date",
        "keywordGroups": body_data[n:m]
    }

    url = "https://openapi.naver.com/v1/datalab/search"


    context = ssl._create_unverified_context()
    request = urllib.request.Request(url)
    request.add_header("X-Naver-Client-Id",client_id)
    request.add_header("X-Naver-Client-Secret",client_secret)
    request.add_header("Content-Type","application/json")
    response = urllib.request.urlopen(request, data=json.dumps(body).encode("utf-8"), context=context)

    rescode = response.getcode()

    if(rescode==200):
        response_body = response.read()
        response_data = response_body.decode('utf-8')
    else:
        print("Error Code:" + rescode)

    result = json.loads(response_data)

    print(result)

    # 엑셀 파일 로드
    workbook = load_workbook('memeNaver.xlsx')

    for i in range(n, m):

        # 결과를 저장할 새로운 엑셀 시트 생성
        result_sheet = workbook.create_sheet(title=str(i))

        # 데이터 입력 받기
        data = []

        print(i-n)

        # 검색량 분석 결과를 엑셀에 저장
        result_sheet.append(['ratio','period'])
        for j in range(len(result['results'][int(i-n)]['data'])):
            new_row = [result['results'][i-n]['data'][j]['ratio'], result['results'][i-n]['data'][j]['period']]
            result_sheet.append(new_row)

    # 변경된 내용을 저장
    workbook.save('memeNaver.xlsx')


for i in range(0, 61):
    getData(i, i+1)